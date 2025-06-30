from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List

class ComplianceReporter:
    def __init__(self):
        pass

    def export_to_excel(self, sections: List[dict], output_path: str) -> str:
        """
        Generate a detailed comparison report directly to Excel.
        'sections' is a list of dicts, each representing a clause/section.
        """
        wb = Workbook()
        summary_sheet = wb.active
        summary_sheet.title = "Summary"

        # ... (move your Excel writing logic here, using 'sections' as input) ...

        wb.save(output_path)
        print(f"✅ Excel report saved to: {output_path}")
        return output_path

    def generate_detailed_comparison(self, sections: List[dict], output_path: str) -> str:
        """
        Generate a detailed Markdown comparison report.
        'sections' is a list of dicts, each representing a clause/section.
        """
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("# DETAILED METER SPECIFICATION COMPLIANCE ANALYSIS\n\n")
            # ... (move your Markdown writing logic here, using 'sections' as input) ...
        print(f"✅ Detailed comparison saved to: {output_path}")
        return output_path

    def format_cell_content(self, text: str, max_length: int = 30) -> str:
        """
        Format cell content for better table readability (for Markdown tables).
        """
        if not text or len(text) <= max_length:
            return text

        words = text.split()
        lines = []
        current_line = []
        current_length = 0

        for word in words:
            if current_length + len(word) + (1 if current_length > 0 else 0) > max_length:
                lines.append(" ".join(current_line))
                current_line = [word]
                current_length = len(word)
            else:
                current_line.append(word)
                current_length += len(word) + (1 if current_length > 0 else 0)

        if current_line:
            lines.append(" ".join(current_line))

        return "<br>".join(lines)

    # Add any other reporting/formatting helpers here
