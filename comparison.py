import os
import re
import json
import ollama
import sqlite3
from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class MeterSpecificationComparison:
    """Tool to compare tender requirements with actual meter specifications"""
    
    def __init__(self, kb_path: str = None):
        """Initialize with path to knowledge base"""
        # Use exact path if not specified
        self.kb_path = kb_path or "C:\\Users\\cyqt2\\Database\\meter_specifications_kb_detailed.json"
        self.model = "qwen2.5-coder:7b"
        
        # Model mapping dictionary to handle variant matching
        self.model_map = {
            "PM8340": "PM83xx",
            "PM8000": "PM8xxx",
            "ION9000": "ION9xxx",
            "PM5000": "PM5xxx",
            # Add more mappings as needed
        }
        
        self.kb_data = self._load_kb()
        print(f"Loaded knowledge base from: {self.kb_path}")
        if self.kb_data:
            print(f"Found {len(self.kb_data)} meter series entries")
    
    def _load_kb(self) -> Dict:
        """Load the meter specifications knowledge base"""
        try:
            with open(self.kb_path, 'r') as f:
                content = f.read()
                # Fix common JSON syntax errors
                content = re.sub(r',\s*]', ']', content)
                content = re.sub(r',\s*}', '}', content)
                return json.loads(content)
        except Exception as e:
            print(f"Error loading knowledge base: {e}")
            return {}
    
    def _extract_sections_from_analysis(self, analysis_path: str) -> List[Dict]:
        """Extract clause sections and selected meters from analysis file"""
        sections = []
        current_section = None
        
        with open(analysis_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            
        for i, line in enumerate(lines):
            # Section header
            if line.startswith('## '):
                if current_section:
                    sections.append(current_section)
                
                # Extract clause ID
                clause_match = re.search(r'Clause (\d+\.\d+\.?\d*)', line)
                if clause_match:
                    current_section = {
                        'clause_id': clause_match.group(1),
                        'requirements': [],
                        'selected_meter': None,
                        'meter_description': '',
                        'line_start': i
                    }
            
            # Requirements
            elif current_section and line.startswith('- ') and i > current_section['line_start'] and 'requirements' not in line.lower():
                if "Selected Meter:" not in lines[i-2]:  # Ensure we're in requirements section
                    current_section['requirements'].append(line[2:].strip())
            
            # Selected meter
            elif current_section and "Selected Meter:" in line:
                meter_match = re.search(r'Selected Meter: ([\w\d]+)', line)
                if meter_match:
                    current_section['selected_meter'] = meter_match.group(1)
            
            # Meter description
            elif current_section and line.startswith('**Description**:'):
                current_section['meter_description'] = line.replace('**Description**:', '').strip()
        
        # Add the last section
        if current_section:
            sections.append(current_section)
        
        return sections
    
    def _find_meter_specs(self, model_number: str) -> Dict:
        """Find specifications for a given meter model with improved matching logic"""
        print(f"Looking for specifications for model: {model_number}")
        
        # Handle PM8000 series directly since it's causing issues
        if model_number.startswith("PM8"):
            if "PM8000_Series" in self.kb_data:
                print(f"✅ Found PM8000 series match")
                series_data = self.kb_data["PM8000_Series"]
                
                # Extract model series digits (PM8340 -> "83")
                if len(model_number) >= 4:
                    model_series = model_number[2:4]
                    print(f"Extracted model series: {model_series}")
                    
                    # Look for matching variant in model_breakdown
                    if "model_breakdown" in series_data:
                        for variant in series_data["model_breakdown"]:
                            variant_name = variant.get("model_name", "")
                            
                            # Check if variant matches series (PM83xx matches PM8340)
                            if model_series in variant_name:
                                print(f"✅ Found PM8000 variant match: {variant_name}")
                                
                                # Create complete specs dictionary
                                specs = {
                                    "model": series_data.get("model", ""),
                                    "summary": series_data.get("summary", ""),
                                    "variant": variant_name,
                                    "variant_features": variant.get("key_differentiator", ""),
                                    "performance_and_accuracy": series_data.get("performance_and_accuracy", []),
                                    "technical_specifications": series_data.get("technical_specifications", {})
                                }
                                
                                return specs
        
        # Extract model family prefix (e.g., 'PM5' from 'PM5350')
        model_prefix = re.match(r'([A-Za-z]+\d+)', model_number)
        if not model_prefix:
            return {}
            
        prefix = model_prefix.group(1)
        print(f"Looking for model family: {prefix}")
        
        # Build the series key (e.g., 'PM5' -> 'PM5000_Series')
        series_key = f"{prefix}000_Series"
        
        if series_key in self.kb_data:
            print(f"✅ Found series: {series_key}")
            series_data = self.kb_data[series_key]
            
            # For other series, use exact model matching
            if "model_breakdown" in series_data:
                for variant in series_data["model_breakdown"]:
                    variant_name = variant.get("model_name", "")
                    
                    # Try different matching strategies
                    if (model_number in variant_name or 
                        model_number == variant_name.split(" ")[0] or
                        model_number.startswith(variant_name.split("x")[0])):
                        
                        print(f"✅ Found model match: {variant_name}")
                        
                        specs = {
                            "model": series_data.get("model", ""),
                            "summary": series_data.get("summary", ""),
                            "variant": variant_name,
                            "variant_features": variant.get("key_differentiator", ""),
                            "performance_and_accuracy": series_data.get("performance_and_accuracy", []),
                            "technical_specifications": series_data.get("technical_specifications", {})
                        }
                        
                        return specs
        
            # If no specific variant found, return series data
            print(f"No specific variant found, using series data")
            return series_data
    
        # Handle iEM series specially
        if model_number.startswith("iEM"):
            if "iEM3000_Series" in self.kb_data:
                print(f"✅ Found iEM series match: iEM3000_Series")
                return self.kb_data["iEM3000_Series"]
        
        # Try direct model lookup in all series as last resort
        for series_key, series_data in self.kb_data.items():
            if "model_breakdown" in series_data:
                for variant in series_data["model_breakdown"]:
                    variant_name = variant.get("model_name", "")
                    if model_number in variant_name:
                        print(f"✅ Found model match in {series_key}: {variant_name}")
                        
                        specs = {
                            "model": series_data.get("model", ""),
                            "summary": series_data.get("summary", ""),
                            "variant": variant_name,
                            "variant_features": variant.get("key_differentiator", ""),
                            "performance_and_accuracy": series_data.get("performance_and_accuracy", []),
                            "technical_specifications": series_data.get("technical_specifications", {})
                        }
                        
                        return specs
        
        print(f"❌ No specifications found for {model_number}")
        print(f"Available series: {list(self.kb_data.keys())}")
        return {}
        
    def _lookup_meter_in_database(self, model_number: str) -> Dict:
        """Look up meter information in the database"""
        db_path = "meters.db"  # Adjust path as needed
        
        # Check if database exists
        if not os.path.exists(db_path):
            print(f"⚠️ Database file not found: {db_path}")
            return {}
        
        try:
            with sqlite3.connect(db_path) as conn:
                cursor = conn.cursor()
                
                # Query for this specific model
                cursor.execute("""
                    SELECT ProductID, ModelNumber, ProductDescription, MID_Certified 
                    FROM Products 
                    WHERE ModelNumber = ?
                    LIMIT 1
                """, (model_number,))
                
                result = cursor.fetchone()
                
                if result:
                    db_info = {
                        "product_id": result[0],
                        "model_number": result[1],
                        "description": result[2],
                        "mid_certified": result[3],
                        "source": "database"
                    }
                    print(f"✅ Found meter in database: {model_number}")
                    return db_info
                
                # Try partial match if exact match fails
                cursor.execute("""
                    SELECT ProductID, ModelNumber, ProductDescription, MID_Certified 
                    FROM Products 
                    WHERE ModelNumber LIKE ?
                    LIMIT 1
                """, (f"{model_number}%",))
                
                result = cursor.fetchone()
                
                if result:
                    db_info = {
                        "product_id": result[0],
                        "model_number": result[1],
                        "description": result[2],
                        "mid_certified": result[3],
                        "source": "database"
                    }
                    print(f"✅ Found similar meter in database: {result[1]}")
                    return db_info
        
        except Exception as e:
            print(f"⚠️ Database lookup error: {e}")
        
        print("ℹ️ No matching meter found in database")
        return {}

    def _compare_requirements_with_specs(self, requirements: List[str], meter_specs: Dict, model_number: str) -> Dict:
        """Use AI to compare tender requirements with meter specifications from both KB and DB"""
        
        # Get additional info from database if available
        db_info = self._lookup_meter_in_database(model_number)
        
        # Extract structured data from the KB specs
        meter_name = meter_specs.get("model", "Unknown")
        summary = meter_specs.get("summary", "")
        variant_name = meter_specs.get("variant", "")
        variant_features = meter_specs.get("variant_features", "")
        
        # Add DB description if available
        db_description = db_info.get("description", "")
        
        # Get performance data (accuracy specs)
        performance_specs = []
        for spec in meter_specs.get("performance_and_accuracy", []):
            if isinstance(spec, dict):
                param = spec.get("parameter", "")
                standard = spec.get("standard", "")
                accuracy = spec.get("class_or_accuracy", "")
                
                if param and accuracy:
                    performance_specs.append(f"{param}: {accuracy} ({standard})")
        
        # Get technical specifications
        tech_specs = []
        tech_data = meter_specs.get("technical_specifications", {})
        
        if isinstance(tech_data, dict):
            for category, details in tech_data.items():
                if isinstance(details, dict):
                    for key, value in details.items():
                        tech_specs.append(f"{category}.{key}: {value}")
                elif isinstance(details, list):
                    tech_specs.append(f"{category}: {', '.join(map(str, details))}")
                else:
                    tech_specs.append(f"{category}: {details}")
        
        # Format spec data for the AI, including both KB and DB data
        formatted_specs = f"""
        MODEL: {meter_name} 
        VARIANT: {variant_name}
        SUMMARY: {summary}
        VARIANT FEATURES: {variant_features}
        
        DATABASE DESCRIPTION: {db_description}
        
        ACCURACY SPECIFICATIONS:
        {chr(10).join(performance_specs)}
        
        TECHNICAL SPECIFICATIONS:
        {chr(10).join(tech_specs)}
        """
        
        # Format requirements
        reqs_formatted = '\n'.join(f"- {req}" for req in requirements)
        
        # Create comparison prompt with improved accuracy class interpretation
        comparison_prompt = f"""
        ## SPECIFICATION COMPARISON TASK
        
        Compare the tender requirements with the meter specifications and determine compliance.
        
        ### TENDER REQUIREMENTS:
        {reqs_formatted}
        
        ### METER SPECIFICATIONS:
        Model: {model_number}
        {formatted_specs}
        
        ### INSTRUCTIONS:
        For each requirement:
        1. Find the relevant specification in the meter data
        2. Determine if the meter complies with the requirement using these rules:
           - Mark as COMPLIANT if the meter meets OR EXCEEDS the requirement
           - For accuracy requirements:
               - "Class X" means accuracy of ±X% (e.g., "Class 0.5" = ±0.5% accuracy)
               - "Class 0.5 S" is better than just "Class 0.5" and means ±0.5% accuracy
               - A lower class number is BETTER (Class 0.2 is better than Class 0.5)
               - When comparing percentages, lower is better (±0.1% is better than ±0.5%)
           - When a spec says "Class 0.5 S (±1%)", the Class rating is the primary metric
        3. Provide justification based on actual meter specs
        4. Note any limitations or potential issues
        
        ACCURACY CLASS INTERPRETATION:
        - "Class 0.5" means ±0.5% accuracy
        - "Class 0.2" means ±0.2% accuracy
        - "Class 0.1" means ±0.1% accuracy
        - If requirement asks for "±0.5% accuracy" and meter is "Class 0.5 S" or better, it COMPLIES
        
        ### OUTPUT FORMAT:
        Return a detailed comparison in JSON format:
        ```json
        {{
          "compliance_analysis": [
            {{
              "requirement": "Voltage accuracy ±0.5%", 
              "spec_value": "Class 0.2 (±0.1% of reading)",
              "complies": true,
              "justification": "The meter's Class 0.2 (±0.1%) EXCEEDS the required ±0.5% accuracy"
            }}
          ],
          "overall_compliance": true,
          "areas_exceeding_requirements": ["Voltage accuracy", "Communication interfaces"],
          "potential_issues": ["No specific certification for revenue metering mentioned"]
        }}
        ```
        
        IMPORTANT: For accuracy requirements, remember that:
        1. "Class X" means ±X% accuracy (Class 0.5 = ±0.5% accuracy)
        2. When a meter has "Class 0.5 S" and requirement is "±0.5%", this COMPLIES
        3. When a meter has "Class 0.5 S (±1%)", focus on the Class 0.5 S part which COMPLIES with "±0.5%" requirement
        """
        
        try:
            response = ollama.chat(
                model=self.model,
                messages=[{"role": "user", "content": comparison_prompt}],
                options={"temperature": 0.2}
            )
            
            ai_content = response['message']['content']
            
            # Extract JSON from AI response
            json_match = re.search(r'\{.*\}', ai_content, re.DOTALL)
            if json_match:
                return json.loads(json_match.group())
            else:
                return {"error": "Could not extract JSON from AI response"}
                
        except Exception as e:
            print(f"Error in AI comparison: {e}")
            return {"error": str(e)}
    
    def generate_detailed_comparison(self, analysis_path: str, override_meter: str = None, per_clause_override: dict = None) -> str:
        """Generate a detailed comparison report with properly formatted tables"""
        sections = self._extract_sections_from_analysis(analysis_path)
        
        if not sections:
            return "No sections found in the analysis file."
        
        # Create output file name
        base_name = os.path.splitext(analysis_path)[0]
        output_path = f"{base_name}_detailed_comparison.md"
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("# DETAILED METER SPECIFICATION COMPLIANCE ANALYSIS\n\n")
            
            for section in sections:
                clause_id = section.get('clause_id', 'Unknown')
                requirements = section.get('requirements', [])
                selected_meter = section.get('selected_meter')
                
                # Per-clause override takes precedence
                if per_clause_override and clause_id in per_clause_override:
                    selected_meter = per_clause_override[clause_id]
                    section['selected_meter'] = selected_meter
                elif override_meter:
                    selected_meter = override_meter
                    section['selected_meter'] = override_meter
                
                f.write(f"## Clause {clause_id} Compliance Analysis\n\n")
                
                if not selected_meter:
                    f.write("No meter selected for this clause.\n\n")
                    continue
                
                f.write(f"### Selected Meter: {selected_meter}\n\n")
                f.write(f"**Description**: {section.get('meter_description', '')}\n\n")
                
                # Get meter specifications from KB
                meter_specs = self._find_meter_specs(selected_meter)
                
                if not meter_specs:
                    f.write(f"⚠️ No detailed specifications found for {selected_meter} in the knowledge base.\n\n")
                    continue
                
                print(f"Analyzing compliance for {selected_meter} against {clause_id} requirements...")
                
                # Compare requirements with specifications
                comparison = self._compare_requirements_with_specs(requirements, meter_specs, selected_meter)
                
                # Write comparison results
                f.write("### Detailed Compliance Analysis\n\n")
                
                if "error" in comparison:
                    f.write(f"⚠️ Error analyzing compliance: {comparison['error']}\n\n")
                    continue
                
                # Get analysis items
                analysis_items = comparison.get("compliance_analysis", [])
                
                if not analysis_items:
                    f.write("No compliance analysis items generated.\n\n")
                    continue
                
                # Create properly formatted table with consistent widths
                f.write("| Requirement | Specification | Status | Justification |\n")
                f.write("|------------|---------------|--------|---------------|\n")
                
                for item in analysis_items:
                    req = item.get("requirement", "Unknown").strip()
                    spec = item.get("spec_value", "Not specified").strip()
                    
                    # Fix the compliance display for better-than-required specs
                    complies = item.get("complies", False)
                    justification = item.get("justification", "").strip()
                    
                    # Check if we have a better-than-required spec that was marked compliant
                    is_better = False
                    if complies:
                        lower_justification = justification.lower()
                        if "exceed" in lower_justification or "better" in lower_justification:
                            is_better = True
                            
                    compliance_status = "✅ Exceeds" if is_better else "✅ Compliant" if complies else "❌ Non-compliant"
                    
                    # Format and write row
                    req = self._format_cell_content(req, 30)
                    spec = self._format_cell_content(spec, 30)
                    justification = self._format_cell_content(justification, 40)
                    
                    f.write(f"| {req} | {spec} | {compliance_status} | {justification} |\n")
                
                f.write("\n")
                
                # Overall assessment
                overall = comparison.get("overall_compliance", False)
                f.write(f"**Overall Compliance**: {'✅ Compliant' if overall else '❌ Non-compliant'}\n\n")
                
                # Areas exceeding requirements
                exceeding = comparison.get("areas_exceeding_requirements", [])
                if exceeding:
                    f.write("**Areas Exceeding Requirements**:\n")
                    for area in exceeding:
                        f.write(f"- {area}\n")
                    f.write("\n")
                
                # Potential issues
                issues = comparison.get("potential_issues", [])
                if issues:
                    f.write("**Potential Compliance Issues**:\n")
                    for issue in issues:
                        f.write(f"- {issue}\n")
                    f.write("\n")
                
                f.write("\n" + "-"*80 + "\n\n")
            
            f.write("\n*Analysis completed using AI-assisted specification comparison*\n")
        
        print(f"✅ Detailed comparison saved to: {output_path}")
        return output_path

    def _format_cell_content(self, text: str, max_length: int = 30) -> str:
        """Format cell content for better table readability
        
        Args:
            text: The text content of the cell
            max_length: Maximum length before adding a line break
            
        Returns:
            Formatted text with line breaks for better table rendering
        """
        if not text or len(text) <= max_length:
            return text
        
        # Split into words
        words = text.split()
        lines = []
        current_line = []
        current_length = 0
        
        for word in words:
            # If adding this word exceeds the max length, start a new line
            if current_length + len(word) + (1 if current_length > 0 else 0) > max_length:
                lines.append(" ".join(current_line))
                current_line = [word]
                current_length = len(word)
            else:
                # Add to current line
                current_line.append(word)
                current_length += len(word) + (1 if current_length > 0 else 0)
        
        # Add the last line if not empty
        if current_line:
            lines.append(" ".join(current_line))
        
        # Join with <br> for markdown line breaks in table cells
        return "<br>".join(lines)
    
    def export_to_excel(self, analysis_path: str, excel_path: str = None, override_meter: str = None, per_clause_override: dict = None) -> str:
        """Generate a detailed comparison report directly to Excel"""
        sections = self._extract_sections_from_analysis(analysis_path)
        
        if not sections:
            print("No sections found in the analysis file.")
            return None
        
        # Create output file name if not specified
        if not excel_path:
            base_name = os.path.splitext(analysis_path)[0]
            excel_path = f"{base_name}_compliance_report.xlsx"
        
        # Create Excel workbook
        wb = Workbook()
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        
        # Add headers to summary sheet
        summary_headers = ['Clause', 'Meter', 'Overall Compliance', 'Compliant Items', 'Non-Compliant Items', 'Key Strengths']
        for i, header in enumerate(summary_headers, 1):
            summary_sheet.cell(row=1, column=i).value = header
            summary_sheet.cell(row=1, column=i).font = Font(bold=True)
        
        # Style for headers
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for col in range(1, len(summary_headers) + 1):
            summary_sheet.cell(row=1, column=col).fill = header_fill
        
        # Add data to summary sheet
        summary_row = 2
        all_sections_compliant = True
        
        for section in sections:
            clause_id = section.get('clause_id', 'Unknown')
            selected_meter = section.get('selected_meter')
            requirements = section.get('requirements', [])
            
            # Per-clause override takes precedence
            if per_clause_override and clause_id in per_clause_override:
                selected_meter = per_clause_override[clause_id]
                section['selected_meter'] = selected_meter
            elif override_meter:
                selected_meter = override_meter
                section['selected_meter'] = override_meter
            
            if not selected_meter:
                continue
            
            # Get meter specifications from KB
            meter_specs = self._find_meter_specs(selected_meter)
            
            if not meter_specs:
                all_sections_compliant = False
                
                # Add to summary as non-compliant due to missing specs
                summary_sheet.cell(row=summary_row, column=1).value = clause_id
                summary_sheet.cell(row=summary_row, column=2).value = selected_meter
                summary_sheet.cell(row=summary_row, column=3).value = "❌ Non-compliant (specs not found)"
                summary_sheet.cell(row=summary_row, column=3).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                summary_row += 1
                continue
            
            # Compare requirements with specifications
            comparison = self._compare_requirements_with_specs(requirements, meter_specs, selected_meter)
            
            if "error" in comparison:
                all_sections_compliant = False
                continue
            
            # Create detail sheet for each section
            sheet_name = f"Clause {clause_id}"
            if len(sheet_name) > 31:  # Excel sheet name length limit
                sheet_name = sheet_name[:31]
            detail_sheet = wb.create_sheet(title=sheet_name)
            
            # Add section info
            detail_sheet.cell(row=1, column=1).value = f"Clause {clause_id} Compliance Analysis"
            detail_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
            detail_sheet.cell(row=2, column=1).value = f"Selected Meter: {selected_meter}"
            detail_sheet.cell(row=2, column=1).font = Font(bold=True)
            detail_sheet.cell(row=3, column=1).value = f"Description: {section.get('meter_description', '')}"
            detail_sheet.cell(row=3, column=1).font = Font(italic=True)
            
            # Get analysis items
            analysis_items = comparison.get("compliance_analysis", [])
            
            # Count compliant and non-compliant items
            compliant = 0
            non_compliant = 0
            for item in analysis_items:
                if item.get("complies", False):
                    compliant += 1
                else:
                    non_compliant += 1
                    all_sections_compliant = False
            
            # Overall compliance for this section
            overall_compliance = comparison.get("overall_compliance", False)
            status_text = "✅ Compliant" if overall_compliance else "❌ Non-compliant"
            detail_sheet.cell(row=4, column=1).value = f"Overall Compliance: {status_text}"
            
            # Add to summary
            summary_sheet.cell(row=summary_row, column=1).value = clause_id
            summary_sheet.cell(row=summary_row, column=2).value = selected_meter
            summary_sheet.cell(row=summary_row, column=3).value = status_text
            summary_sheet.cell(row=summary_row, column=4).value = compliant
            summary_sheet.cell(row=summary_row, column=5).value = non_compliant
            
            # Get areas exceeding requirements
            exceeding_areas = comparison.get("areas_exceeding_requirements", [])
            summary_sheet.cell(row=summary_row, column=6).value = ', '.join(exceeding_areas)
            
            # Color code compliance
            if overall_compliance:
                summary_sheet.cell(row=summary_row, column=3).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                summary_sheet.cell(row=summary_row, column=3).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            summary_row += 1
            
            # Add table headers
            headers = ['Requirement', 'Specification', 'Status', 'Justification']
            for i, header in enumerate(headers, 1):
                detail_sheet.cell(row=6, column=i).value = header
                detail_sheet.cell(row=6, column=i).font = Font(bold=True)
                detail_sheet.cell(row=6, column=i).fill = header_fill
            
            # Add table data
            row_num = 7
            for item in analysis_items:
                # Format requirement
                req = item.get("requirement", "Unknown")
                detail_sheet.cell(row=row_num, column=1).value = req
                
                # Format specification
                spec = item.get("spec_value", "Not specified")
                detail_sheet.cell(row=row_num, column=2).value = spec
                
                # Format status with special handling for "exceeds"
                complies = item.get("complies", False)
                justification = item.get("justification", "")
                
                # Check if we have a better-than-required spec
                is_better = False
                if complies:
                    lower_justification = justification.lower()
                    if "exceed" in lower_justification or "better" in lower_justification:
                        is_better = True
                
                status_text = "✅ Exceeds" if is_better else "✅ Compliant" if complies else "❌ Non-compliant"
                detail_sheet.cell(row=row_num, column=3).value = status_text
                
                # Format justification
                detail_sheet.cell(row=row_num, column=4).value = justification
                
                # Color code status
                if complies:
                    detail_sheet.cell(row=row_num, column=3).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    detail_sheet.cell(row=row_num, column=3).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                row_num += 1
            
            # Add exceeding areas
            if exceeding_areas:
                row_num += 1
                detail_sheet.cell(row=row_num, column=1).value = "Areas Exceeding Requirements:"
                detail_sheet.cell(row=row_num, column=1).font = Font(bold=True)
                for area in exceeding_areas:
                    row_num += 1
                    detail_sheet.cell(row=row_num, column=1).value = f"- {area}"
            
            # Add issues
            issues = comparison.get("potential_issues", [])
            if issues:
                row_num += 1
                detail_sheet.cell(row=row_num, column=1).value = "Potential Compliance Issues:"
                detail_sheet.cell(row=row_num, column=1).font = Font(bold=True)
                for issue in issues:
                    row_num += 1
                    detail_sheet.cell(row=row_num, column=1).value = f"- {issue}"
            
            # Auto-adjust column widths
            for col in range(1, 5):
                max_length = 0
                for row in range(6, row_num + 1):
                    cell_value = detail_sheet.cell(row=row, column=col).value
                    if cell_value:
                        text_length = len(str(cell_value).split('\n')[0])  # First line length
                        max_length = max(max_length, min(text_length, 80))  # Cap at 80
                
                # Set width with some padding
                adjusted_width = max_length + 2
                detail_sheet.column_dimensions[get_column_letter(col)].width = adjusted_width
            
            # Add borders and text wrapping
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
                               
            for row in range(6, row_num):
                for col in range(1, 5):
                    detail_sheet.cell(row=row, column=col).border = thin_border
                    detail_sheet.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    
    # Auto-adjust summary column widths
        for col in range(1, len(summary_headers) + 1):
            max_length = len(str(summary_headers[col-1]))
            for row in range(2, summary_row):
                cell_value = summary_sheet.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set width with some padding
            summary_sheet.column_dimensions[get_column_letter(col)].width = max_length + 2
        
        # Add borders to summary
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
                            
        for row in range(1, summary_row):
            for col in range(1, len(summary_headers) + 1):
                summary_sheet.cell(row=row, column=col).border = thin_border
        
        # Save workbook
        wb.save(excel_path)
        print(f"✅ Excel report saved to: {excel_path}")
        return excel_path
    

if __name__ == "__main__":
    print("\nMeter Specification Compliance Tool\n" + "="*35)
    
    # Get analysis file path
    analysis_file = input("Enter path to meter analysis file (.txt): ").strip('"\'')
    
    if not os.path.exists(analysis_file):
        print(f"Error: Analysis file '{analysis_file}' not found!")
        exit(1)
    
    # Create comparator
    comparator = MeterSpecificationComparison()
    
    # Extract sections first
    sections = comparator._extract_sections_from_analysis(analysis_file)
    if not sections:
        print("No sections found in the analysis file.")
        exit(1)

    # Collect per-clause overrides
    per_clause_override = {}
    print("\nReview recommended meters for each clause. Press Enter to accept or type an alternate model number.\n")
    for section in sections:
        clause_id = section.get('clause_id', 'Unknown')
        recommended = section.get('selected_meter', 'None')
        print(f"Clause {clause_id} - {recommended}")
        alt = input("Alternate Meter (Enter to default): ").strip()
        if alt:
            per_clause_override[clause_id] = alt
        else:
            per_clause_override[clause_id] = recommended

    # Ask for output format preference
    output_format = input("\nSelect output format:\n1. Markdown\n2. Excel\n3. Both\nEnter choice (1-3): ").strip()
    
    # Generate outputs based on preference
    md_output = None
    excel_output = None

    # Pass per-clause overrides to the report functions
    def get_override(clause_id):
        return per_clause_override.get(clause_id)

    if output_format in ["1", "3"]:
        # Patch: generate_detailed_comparison now takes a dict of clause_id -> meter
        md_output = comparator.generate_detailed_comparison(
            analysis_file,
            override_meter=None,
            per_clause_override=per_clause_override
        )
        print(f"Markdown report generated: {md_output}")
    
    if output_format in ["2", "3"]:
        excel_output = comparator.export_to_excel(
            analysis_file,
            override_meter=None,
            per_clause_override=per_clause_override
        )
        print(f"Excel report generated: {excel_output}")
    
    print("\nAnalysis complete!")