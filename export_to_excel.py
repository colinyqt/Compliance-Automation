import re
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def extract_tables_from_markdown(markdown_file):
    """Extract tables and section info from markdown file"""
    with open(markdown_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Split by clauses
    clause_sections = re.split(r'## Clause (\d+\.\d+\.\d+) Compliance Analysis', content)[1:]
    
    results = []
    for i in range(0, len(clause_sections), 2):
        if i+1 >= len(clause_sections):
            break
            
        clause_id = clause_sections[i].strip()
        section_content = clause_sections[i+1].strip()
        
        # Extract meter info
        meter_match = re.search(r'### Selected Meter: ([^\n]+)', section_content)
        meter = meter_match.group(1) if meter_match else "Unknown"
        
        desc_match = re.search(r'\*\*Description\*\*: ([^\n]+)', section_content)
        description = desc_match.group(1) if desc_match else ""
        
        # Extract table content
        table_match = re.search(r'\| Requirement \| Specification \| Status \| Justification \|\s*\|\s*-+\s*\|\s*-+\s*\|\s*-+\s*\|\s*-+\s*\|(.*?)(?:\n\n\*\*Overall|\n\n\n|$)', section_content, re.DOTALL)
        
        if not table_match:
            continue
            
        table_content = table_match.group(1).strip()
        
        # Extract rows
        rows = []
        for row in table_content.split('\n'):
            if row.strip() and '|' in row:
                cells = [cell.strip().replace('<br>', '\n') for cell in row.split('|')[1:-1]]
                if len(cells) == 4:  # Ensure we have all columns
                    rows.append({
                        'Requirement': cells[0],
                        'Specification': cells[1],
                        'Status': cells[2],
                        'Justification': cells[3],
                        'Clause': clause_id,
                        'Meter': meter,
                        'Description': description
                    })
        
        # Get overall compliance
        overall_match = re.search(r'\*\*Overall Compliance\*\*: ([^\n]+)', section_content)
        overall = overall_match.group(1) if overall_match else "Unknown"
        
        # Get areas exceeding requirements
        exceeding_areas = []
        exceeding_section = re.search(r'\*\*Areas Exceeding Requirements\*\*:\s*((?:- [^\n]+\n)+)', section_content)
        if exceeding_section:
            for line in exceeding_section.group(1).split('\n'):
                if line.strip().startswith('- '):
                    exceeding_areas.append(line.strip()[2:])
        
        # Get potential issues
        issues = []
        issues_section = re.search(r'\*\*Potential Compliance Issues\*\*:\s*((?:- [^\n]+\n)+)', section_content)
        if issues_section:
            for line in issues_section.group(1).split('\n'):
                if line.strip().startswith('- '):
                    issues.append(line.strip()[2:])
        
        results.append({
            'clause_id': clause_id,
            'meter': meter,
            'description': description,
            'rows': rows,
            'overall': overall,
            'exceeding_areas': exceeding_areas,
            'issues': issues
        })
    
    return results

def create_excel_report(markdown_file, excel_file=None):
    """Create Excel report from markdown file"""
    if not excel_file:
        excel_file = os.path.splitext(markdown_file)[0] + '.xlsx'
    
    # Extract data
    sections = extract_tables_from_markdown(markdown_file)
    if not sections:
        print("No compliance data found in file")
        return
    
    # Create Excel workbook
    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    
    # Add headers to summary sheet
    summary_headers = ['Clause', 'Meter', 'Overall Compliance', 'Compliant Items', 'Non-Compliant Items', 'Areas Exceeding Requirements']
    for i, header in enumerate(summary_headers, 1):
        summary_sheet.cell(row=1, column=i).value = header
        summary_sheet.cell(row=1, column=i).font = Font(bold=True)
    
    # Style for headers
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for col in range(1, len(summary_headers) + 1):
        summary_sheet.cell(row=1, column=col).fill = header_fill
    
    # Add data to summary sheet
    summary_row = 2
    for section in sections:
        # Count compliant and non-compliant items
        compliant = 0
        non_compliant = 0
        for row in section['rows']:
            if '✅' in row['Status']:
                compliant += 1
            elif '❌' in row['Status']:
                non_compliant += 1
        
        # Add to summary
        summary_sheet.cell(row=summary_row, column=1).value = section['clause_id']
        summary_sheet.cell(row=summary_row, column=2).value = section['meter']
        summary_sheet.cell(row=summary_row, column=3).value = section['overall']
        summary_sheet.cell(row=summary_row, column=4).value = compliant
        summary_sheet.cell(row=summary_row, column=5).value = non_compliant
        summary_sheet.cell(row=summary_row, column=6).value = ', '.join(section['exceeding_areas'])
        
        # Color code compliance
        if '✅' in section['overall']:
            summary_sheet.cell(row=summary_row, column=3).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif '❌' in section['overall']:
            summary_sheet.cell(row=summary_row, column=3).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        summary_row += 1
        
        # Create detail sheet for each section
        sheet_name = f"Clause {section['clause_id']}"
        if len(sheet_name) > 31:  # Excel sheet name length limit
            sheet_name = sheet_name[:31]
        detail_sheet = wb.create_sheet(title=sheet_name)
        
        # Add section info
        detail_sheet.cell(row=1, column=1).value = f"Clause {section['clause_id']} Compliance Analysis"
        detail_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        detail_sheet.cell(row=2, column=1).value = f"Selected Meter: {section['meter']}"
        detail_sheet.cell(row=2, column=1).font = Font(bold=True)
        detail_sheet.cell(row=3, column=1).value = f"Description: {section['description']}"
        detail_sheet.cell(row=3, column=1).font = Font(italic=True)
        detail_sheet.cell(row=4, column=1).value = f"Overall Compliance: {section['overall']}"
        
        # Add table headers
        headers = ['Requirement', 'Specification', 'Status', 'Justification']
        for i, header in enumerate(headers, 1):
            detail_sheet.cell(row=6, column=i).value = header
            detail_sheet.cell(row=6, column=i).font = Font(bold=True)
            detail_sheet.cell(row=6, column=i).fill = header_fill
        
        # Add table data
        row_num = 7
        for row_data in section['rows']:
            detail_sheet.cell(row=row_num, column=1).value = row_data['Requirement']
            detail_sheet.cell(row=row_num, column=2).value = row_data['Specification']
            detail_sheet.cell(row=row_num, column=3).value = row_data['Status']
            detail_sheet.cell(row=row_num, column=4).value = row_data['Justification']
            
            # Color code status
            if '✅' in row_data['Status']:
                detail_sheet.cell(row=row_num, column=3).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif '❌' in row_data['Status']:
                detail_sheet.cell(row=row_num, column=3).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row_num += 1
        
        # Add exceeding areas
        if section['exceeding_areas']:
            row_num += 1
            detail_sheet.cell(row=row_num, column=1).value = "Areas Exceeding Requirements:"
            detail_sheet.cell(row=row_num, column=1).font = Font(bold=True)
            for area in section['exceeding_areas']:
                row_num += 1
                detail_sheet.cell(row=row_num, column=1).value = f"- {area}"
        
        # Add issues
        if section['issues']:
            row_num += 1
            detail_sheet.cell(row=row_num, column=1).value = "Potential Compliance Issues:"
            detail_sheet.cell(row=row_num, column=1).font = Font(bold=True)
            for issue in section['issues']:
                row_num += 1
                detail_sheet.cell(row=row_num, column=1).value = f"- {issue}"
        
        # Auto-adjust column widths
        for col in range(1, 5):
            max_length = 0
            for row in range(6, row_num + 1):
                cell_value = detail_sheet.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set width with some padding
            adjusted_width = min(max_length + 2, 80)  # Cap at 80 to avoid too wide columns
            detail_sheet.column_dimensions[get_column_letter(col)].width = adjusted_width
        
        # Add more formatting
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
                           
        for row in range(6, row_num):
            for col in range(1, 5):
                detail_sheet.cell(row=row, column=col).border = thin_border
                detail_sheet.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    
    # Auto-adjust summary column widths
    for col in range(1, len(summary_headers) + 1):
        max_length = len(str(summary_headers[col-1]))
        for row in range(2, summary_row):
            cell_value = summary_sheet.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        # Set width with some padding
        summary_sheet.column_dimensions[get_column_letter(col)].width = max_length + 2
    
    # Add borders to summary
    for row in range(1, summary_row):
        for col in range(1, len(summary_headers) + 1):
            summary_sheet.cell(row=row, column=col).border = thin_border
    
    # Save workbook
    wb.save(excel_file)
    print(f"Excel report saved to: {excel_file}")
    return excel_file

if __name__ == "__main__":
    markdown_file = input("Enter path to markdown compliance report: ").strip('"\'')
    if not os.path.exists(markdown_file):
        print(f"Error: File '{markdown_file}' not found!")
    else:
        excel_file = create_excel_report(markdown_file)
        print(f"\nExport complete! Excel file saved to: {excel_file}")