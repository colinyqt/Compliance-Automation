import os
import re
import json
import ollama
import sqlite3
import sys
import time
from typing import Dict, List, Any
from openpyxl import Workbook, DEBUG
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class MeterSpecificationComparison:
    """Tool to compare tender requirements with actual meter specifications"""

    def __init__(self, db_path="overhaul\databases\meters.db"):
        """Initialize with the SQLite database"""
        self.db_path = r"C:\Users\cyqt2\Database\overhaul\databases\meters.db"
        self.model = "qwen2.5-coder:7b"
        self._validate_database()

    def _validate_database(self):
        """Ensure database exists and has expected structure"""
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                tables = [row[0] for row in cursor.fetchall()]
                if 'Meters' not in tables:
                    raise ValueError("Meters table not found in database")
                print(f"✅ Database validated: {len(tables)} tables found")
        except Exception as e:
            raise ValueError(f"Database validation failed: {e}")

    def _find_meter_specs(self, model_number: str) -> dict:
        """Find meter specifications using the SQLite database"""
        if not os.path.exists(self.db_path):
            print(f"❌ Database file not found: {self.db_path}")
            return {}
        
        model_number = model_number.strip()
        print(f"🔍 Searching for meter: {model_number}")
        
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                
                # Try multiple matching strategies based on your schema
                search_queries = [
                    # Exact matches
                    ("EXACT model_name", "SELECT * FROM Meters WHERE model_name = ? LIMIT 1", (model_number,)),
                    ("EXACT device_short_name", "SELECT * FROM Meters WHERE device_short_name = ? LIMIT 1", (model_number,)),
                    ("EXACT series_name", "SELECT * FROM Meters WHERE series_name = ? LIMIT 1", (model_number,)),
                    
                    # Case-insensitive matches
                    ("CASE-INSENSITIVE model_name", "SELECT * FROM Meters WHERE UPPER(model_name) = UPPER(?) LIMIT 1", (model_number,)),
                    ("CASE-INSENSITIVE device_short_name", "SELECT * FROM Meters WHERE UPPER(device_short_name) = UPPER(?) LIMIT 1", (model_number,)),
                    
                    # Partial matches
                    ("PARTIAL model_name", "SELECT * FROM Meters WHERE model_name LIKE ? LIMIT 1", (f"%{model_number}%",)),
                    ("PARTIAL device_short_name", "SELECT * FROM Meters WHERE device_short_name LIKE ? LIMIT 1", (f"%{model_number}%",)),
                    ("PARTIAL product_name", "SELECT * FROM Meters WHERE product_name LIKE ? LIMIT 1", (f"%{model_number}%",)),
                ]
                
                meter_row = None
                matched_by = None
                
                for search_name, query, params in search_queries:
                    cursor.execute(query, params)
                    meter_row = cursor.fetchone()
                    if meter_row:
                        matched_by = search_name
                        break
                
                if not meter_row:
                    print(f"❌ No meter found matching: {model_number}")
                    return {}

                meter_id = meter_row["id"]
                specs = dict(meter_row)
                print(f"✅ Found meter: {specs.get('model_name', model_number)} (ID: {meter_id}) via {matched_by}")

                # Fetch all related specifications using your actual table structure
                
                # Applications
                cursor.execute("SELECT application FROM DeviceApplications WHERE meter_id = ?", (meter_id,))
                specs["applications"] = [row[0] for row in cursor.fetchall()]
                
                # Power Quality Features
                cursor.execute("SELECT analysis_feature FROM PowerQualityAnalysis WHERE meter_id = ?", (meter_id,))
                specs["power_quality_features"] = [row[0] for row in cursor.fetchall()]
                
                # Measurements
                cursor.execute("SELECT measurement_type FROM Measurements WHERE meter_id = ?", (meter_id,))
                specs["measurements"] = [row[0] for row in cursor.fetchall()]
                
                # Accuracy Classes
                cursor.execute("SELECT accuracy_class FROM AccuracyClasses WHERE meter_id = ?", (meter_id,))
                specs["accuracy_classes"] = [row[0] for row in cursor.fetchall()]
                
                # Measurement Accuracy (parameter -> accuracy mapping)
                cursor.execute("SELECT parameter, accuracy FROM MeasurementAccuracy WHERE meter_id = ?", (meter_id,))
                specs["measurement_accuracy"] = {row[0]: row[1] for row in cursor.fetchall()}
                
                # Communication Protocols (protocol -> support mapping)
                cursor.execute("SELECT protocol, support FROM CommunicationProtocols WHERE meter_id = ?", (meter_id,))
                specs["communication_protocols"] = {row[0]: row[1] for row in cursor.fetchall()}
                
                # Data Recording
                cursor.execute("SELECT recording_type FROM DataRecordings WHERE meter_id = ?", (meter_id,))
                specs["data_recording"] = [row[0] for row in cursor.fetchall()]
                
                # Certifications
                cursor.execute("SELECT certification FROM Certifications WHERE meter_id = ?", (meter_id,))
                specs["certifications"] = [row[0] for row in cursor.fetchall()]
                
                # Inputs/Outputs
                cursor.execute("SELECT io_type, description FROM InputsOutputs WHERE meter_id = ?", (meter_id,))
                specs["inputs_outputs"] = [
                    {"type": row[0], "description": row[1]} 
                    for row in cursor.fetchall()
                ]
                
                # Count the specification categories loaded
                spec_count = sum([
                    len(specs.get("applications", [])),
                    len(specs.get("power_quality_features", [])),
                    len(specs.get("measurements", [])),
                    len(specs.get("accuracy_classes", [])),
                    len(specs.get("measurement_accuracy", {})),
                    len(specs.get("communication_protocols", {})),
                    len(specs.get("data_recording", [])),
                    len(specs.get("certifications", [])),
                    len(specs.get("inputs_outputs", [])),
                ])
                
                print(f"📊 Loaded meter specs with {spec_count} detailed specifications")
                return specs
                
        except Exception as e:
            print(f"❌ Error querying database: {e}")
            import traceback
            traceback.print_exc()
            return {}

    def _extract_sections_from_analysis(self, analysis_path: str) -> list:
        """Extract clause sections and selected meters from the analysis output format."""
        print(f"Extracting sections from {analysis_path}...")
        sections = []
        
        try:
            with open(analysis_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
        except Exception as e:
            print(f"Error reading analysis file: {e}")
            return sections

        i = 0
        while i < len(lines):
            line = lines[i]
            
            # Find start of a requirement block
            req_match = re.match(r'✨ Processing requirement \d+/\d+: ([\d\.]+)\.\.\.', line)
            
            if req_match:
                clause_id = req_match.group(1)
                print(f"Found clause: {clause_id}")
                
                section = {
                    'clause_id': clause_id,
                    'requirements': [],
                    'selected_meter': None,
                    'meter_description': ''
                }
                
                # Find the meter type
                i += 1
                while i < len(lines) and not lines[i].strip().startswith("📝 Type:"):
                    i += 1
                
                if i < len(lines) and lines[i].strip().startswith("📝 Type:"):
                    meter_type = lines[i].strip().replace("📝 Type:", "").strip()
                    section['meter_type'] = meter_type
                
                # Find the start of the specifications block
                while i < len(lines) and not lines[i].strip().startswith("📝 Specifications:"):
                    i += 1
                    
                if i < len(lines) and lines[i].strip().startswith("📝 Specifications:"):
                    i += 1  # Skip the "📝 Specifications:" line
                    print(f"Extracting specifications for clause {clause_id}...")
                    
                    # Collect all lines until the next block
                    while i < len(lines):
                        line_strip = lines[i].strip()
                        # Stop if we hit the next block
                        if line_strip.startswith("🏆 Top 3 Best-fit meters:") or \
                           line_strip.startswith("✨ Processing requirement") or \
                           line_strip.startswith("📊 Analysis complete!"):

                            break
                        # Add all non-empty lines as requirements
                        if line_strip:
                            section['requirements'].append(line_strip)
                        i += 1
                        
                # Find the top recommended meter (use the first one as "selected")
                while i < len(lines) and not lines[i].strip().startswith("🏆 Top 3 Best-fit meters:"):
                    i += 1
                    
                if i < len(lines) and lines[i].strip().startswith("🏆 Top 3 Best-fit meters:"):
                    i += 1  # Skip the header line
                    
                    # Look for the first meter (marked with "1.")
                    while i < len(lines):
                        line_strip = lines[i].strip()
                        
                        # Stop if we hit the next processing requirement or end
                        if line_strip.startswith("✨ Processing requirement") or \
                           line_strip.startswith("📊 Analysis complete!"):
                            break
                        
                        # Look for the first recommended meter
                        meter_match = re.match(r'\s*1\.\s+(.+)', line_strip)
                        if meter_match:
                            meter_name = meter_match.group(1).strip()
                            section['selected_meter'] = meter_name
                            print(f"Selected meter for {clause_id}: {meter_name}")
                            
                            # Try to extract description from subsequent lines
                            i += 1
                            description_lines = []
                            while i < len(lines):
                                desc_line = lines[i].strip()
                                if desc_line.startswith("Description:"):
                                    description_lines.append(desc_line.replace("Description:", "").strip())
                                elif desc_line.startswith("2.") or desc_line.startswith("✨") or desc_line.startswith("📊"):
                                    break
                                elif desc_line and not desc_line.startswith("Reason:") and not desc_line.startswith("Score:"):
                                    description_lines.append(desc_line)
                                i += 1
                            
                            # Update the meter description to match the selected meter (not available here)
                            section['meter_description'] = ''
                            break
                        
                        i += 1
                
                if section['requirements']:
                    sections.append(section)
                    print(f"Added section {clause_id} with {len(section['requirements'])} requirements")
            
            i += 1
        
        print(f"Extracted {len(sections)} sections from analysis file")
        return sections

    def _compare_requirements_with_specs(self, requirements: List[str], meter_specs: Dict, model_number: str) -> Dict:
        """Use AI to compare tender requirements with meter specifications"""
        
        # If no requirements or specs, return empty result
        if not requirements:
            print(f"No requirements to analyze for {model_number}")
            return {"error": "No requirements to analyze"}
            
        if not meter_specs:
            print(f"No specifications found for {model_number}")
            return {"error": f"No specifications found for {model_number}"}
            
        print(f"\nComparing {len(requirements)} requirements against {model_number} specifications...")
        
        # If we have too many requirements, chunk them
        MAX_REQUIREMENTS_PER_CHUNK = 10
        
        if len(requirements) > MAX_REQUIREMENTS_PER_CHUNK:
            print(f"🔄 Large requirement set detected. Processing in chunks of {MAX_REQUIREMENTS_PER_CHUNK}...")
            return self._compare_requirements_chunked(requirements, meter_specs, model_number, MAX_REQUIREMENTS_PER_CHUNK)
        
        # Format meter specs for the AI prompt
        meter_name = meter_specs.get("model_name", model_number)
        specs_formatted = self._format_meter_specs_for_prompt(meter_specs)
        reqs_formatted = '\n'.join(f"{i+1}. {req}" for i, req in enumerate(requirements))

        # Build a MUCH more explicit prompt with accuracy class education
        comparison_prompt = f"""
### CRITICAL COMPLIANCE RULES - READ CAREFULLY

**ACCURACY CLASS HIERARCHY (SMALLER NUMBER = BETTER ACCURACY)**:
- IEC 61557-12: Class 0.02 > Class 0.05 > Class 0.1 > Class 0.2 > Class 0.5 > Class 1
- IEC 62053-22: Class 0.1S > Class 0.2S > Class 0.5S > Class 1 > Class 2
- IEC 62053-24: Class 0.5S > Class 1 > Class 2 > Class 3

**PERCENTAGE ACCURACY (SMALLER = BETTER)**:
- ±0.02% > ±0.05% > ±0.1% > ±0.2% > ±0.5% > ±1.0%

**COMPLIANCE LOGIC**:
- If meter spec is BETTER than requirement → Mark as ✅ **COMPLIANT** (and note "exceeds")
- If meter spec EQUALS requirement → Mark as ✅ **COMPLIANT**  
- If meter spec is WORSE than requirement → Mark as ❌ **NON-COMPLIANT**
- If spec not found → Mark as ❌ **NON-COMPLIANT**

**EXAMPLES TO LEARN FROM**:
- Requirement: "±0.2%" | Meter: "Class 0.1" → ✅ COMPLIANT (0.1% is better than 0.2%)
- Requirement: "Class 0.2S" | Meter: "Class 0.1S" → ✅ COMPLIANT (0.1S is better than 0.2S)  
- Requirement: "Class 2" | Meter: "Class 0.5S" → ✅ COMPLIANT (0.5S is much better than Class 2)
- Requirement: "±0.1%" | Meter: "±0.2%" → ❌ NON-COMPLIANT (0.2% is worse than 0.1%)

### REQUIREMENTS TO ANALYZE (ALL {len(requirements)} ITEMS)
{reqs_formatted}

### METER SPECIFICATIONS
Model: {model_number}
{specs_formatted}

### TASK
Analyze each requirement using the compliance rules above. For each requirement:
1. Find the relevant meter specification
2. Compare using the correct hierarchy (smaller accuracy numbers are BETTER)
3. Mark as COMPLIANT if meter meets OR EXCEEDS the requirement
4. In justification, mention if meter "exceeds requirement" when applicable

### REQUIRED JSON OUTPUT
{{
  "compliance_analysis": [
    {{"requirement": "requirement 1 text", "spec_value": "meter specification", "complies": true/false, "justification": "detailed reason with 'exceeds requirement' if applicable"}},
    {{"requirement": "requirement 2 text", "spec_value": "meter specification", "complies": true/false, "justification": "detailed reason with 'exceeds requirement' if applicable"}},
    ... continue for all {len(requirements)} requirements
  ],
  "overall_compliance": true/false,
  "areas_exceeding_requirements": ["list any areas where meter exceeds requirements"],
  "potential_issues": []
}}

**CRITICAL**: Apply the accuracy hierarchy correctly. A meter with Class 0.1 accuracy EXCEEDS a requirement for Class 0.2 accuracy.
"""
    
        # Track start time
        start_time = time.time()
        print(f"Sending corrected compliance request to Qwen for {model_number}...")
        
        try:
            response = ollama.chat(
                model=self.model,
                messages=[{"role": "user", "content": comparison_prompt}],
                options={
                    "temperature": 0.05,    # Even lower temperature for consistency
                    "timeout": 180.0,
                    "num_ctx": 16384,
                    "num_predict": 8192
                }
            )
            
            elapsed = time.time() - start_time
            print(f"✓ Received response in {elapsed:.1f}s for {model_number}")
            
            ai_content = response['message']['content']
            
            # Save the corrected response for debugging
            debug_file = f"debug_corrected_response_{model_number.replace(' ', '_').replace('/', '_')}.txt"
            with open(debug_file, 'w', encoding='utf-8') as f:
                f.write(f"CORRECTED COMPLIANCE ANALYSIS\n")
                f.write(f"Model: {model_number}\n")
                f.write(f"Requirements count: {len(requirements)}\n\n")
                f.write("CORRECTED AI RESPONSE:\n")
                f.write(ai_content)
            
            # Use the robust JSON extraction
            result = self._extract_and_repair_json(ai_content)
            
            # POST-PROCESS: Double-check the compliance logic in case AI still gets it wrong
            result = self._post_process_compliance_logic(result, requirements)
            
            return result

        except Exception as e:
            elapsed = time.time() - start_time
            print(f"❌ Error after {elapsed:.1f}s: {e}")
            return {"error": f"AI comparison failed: {str(e)}"}

    def _compare_requirements_chunked(self, requirements: List[str], meter_specs: Dict, model_number: str, chunk_size: int) -> Dict:
        """Compare requirements in chunks to handle large requirement sets"""
        
        print(f"🔄 Processing {len(requirements)} requirements in chunks of {chunk_size}")
        
        # Split requirements into chunks
        chunks = [requirements[i:i + chunk_size] for i in range(0, len(requirements), chunk_size)]
        print(f"📦 Created {len(chunks)} chunks")
        
        all_compliance_items = []
        chunk_results = []
        
        for i, chunk in enumerate(chunks, 1):
            print(f"\n📦 Processing chunk {i}/{len(chunks)} ({len(chunk)} requirements)...")
            
            # Process this chunk
            chunk_result = self._compare_requirements_with_specs_single_chunk(chunk, meter_specs, model_number, i)
            
            if "error" in chunk_result:
                print(f"❌ Error in chunk {i}: {chunk_result['error']}")
                continue
            
            chunk_compliance = chunk_result.get("compliance_analysis", [])
            all_compliance_items.extend(chunk_compliance)
            chunk_results.append(chunk_result)
            
            print(f"✅ Chunk {i} completed: {len(chunk_compliance)} items analyzed")
            
            # Small delay between chunks to avoid overwhelming the AI
            if i < len(chunks):
                time.sleep(2)
        
        print(f"\n📊 Chunked processing complete: {len(all_compliance_items)} total items analyzed")
        
        # Combine results
        combined_result = {
            "compliance_analysis": all_compliance_items,
            "overall_compliance": True,
            "areas_exceeding_requirements": [],
            "potential_issues": []
        }
        
        # Calculate overall compliance
        compliant_count = sum(1 for item in all_compliance_items if item.get("complies", False))
        total_count = len(all_compliance_items)
        compliance_rate = compliant_count / total_count if total_count > 0 else 0
        
        combined_result["overall_compliance"] = compliance_rate >= 0.8  # 80% compliance threshold
        
        # Combine areas exceeding requirements - FIX: Handle string lists properly
        all_exceeding = []
        for chunk_result in chunk_results:
            exceeding_areas = chunk_result.get("areas_exceeding_requirements", [])
            if isinstance(exceeding_areas, list):
                all_exceeding.extend(exceeding_areas)
        
        # Remove duplicates by converting to set and back (only works with strings)
        combined_result["areas_exceeding_requirements"] = self._safe_remove_duplicates(all_exceeding)
        
        # Combine potential issues - FIX: Handle string lists properly
        all_issues = []
        for chunk_result in chunk_results:
            issues = chunk_result.get("potential_issues", [])
            if isinstance(issues, list):
                all_issues.extend(issues)
    
        # Remove duplicates by converting to set and back (only works with strings)
        combined_result["potential_issues"] = self._safe_remove_duplicates(all_issues)
    
        return combined_result

    def _compare_requirements_with_specs_single_chunk(self, requirements: List[str], meter_specs: Dict, model_number: str, chunk_num: int) -> Dict:
        """Process a single chunk of requirements"""
        
        # Format meter specs for the AI prompt
        specs_formatted = self._format_meter_specs_for_prompt(meter_specs)
        reqs_formatted = '\n'.join(f"{i+1}. {req}" for i, req in enumerate(requirements))

        # Build prompt for this chunk
        comparison_prompt = f"""
### TASK
Analyze ALL {len(requirements)} requirements in this chunk against meter specifications.

### REQUIREMENTS TO ANALYZE (Chunk {chunk_num})
{reqs_formatted}

### METER SPECIFICATIONS
Model: {model_number}
{specs_formatted}

### OUTPUT
Valid JSON with analysis for ALL {len(requirements)} requirements:

{{
  "compliance_analysis": [
    {{"requirement": "requirement text", "spec_value": "meter spec", "complies": true/false, "justification": "reason"}}
  ],
  "overall_compliance": true/false,
  "areas_exceeding_requirements": [],
  "potential_issues": []
}}

Analyze ALL {len(requirements)} requirements. Do not skip any.
"""
    
        try:
            response = ollama.chat(
                model=self.model,
                messages=[{"role": "user", "content": comparison_prompt}],
                options={
                    "temperature": 0.1,
                    "timeout": 120.0,
                    "num_ctx": 8192,
                    "num_predict": 4096
                }
            )
            
            ai_content = response['message']['content']
            result = self._extract_and_repair_json(ai_content)
            
            return result
            
        except Exception as e:
            return {"error": f"Chunk {chunk_num} failed: {str(e)}"}

    def _extract_missing_compliance_items(self, ai_content: str, existing_items: list, original_requirements: list) -> list:
        """Try to extract any missing compliance items from the AI response"""
        
        print(f"🔧 Attempting to recover missing compliance items...")
        
        # Get requirements that are already analyzed (normalize for comparison)
        analyzed_reqs = set()
        for item in existing_items:
            req_text = item.get("requirement", "").strip()
            if req_text:
                # Remove numbering and get first 50 chars for comparison
                normalized = re.sub(r'^\d+\.\s*', '', req_text).strip()[:50]
                analyzed_reqs.add(normalized)
        
        print(f"🔧 Already analyzed {len(analyzed_reqs)} requirements")
        
        # Find missing requirements
        missing_reqs = []
        for i, req in enumerate(original_requirements):
            normalized = req.strip()[:50]
            if normalized not in analyzed_reqs:
                missing_reqs.append((i, req))
        
        print(f"🔧 Found {len(missing_reqs)} potentially missing requirements")
        
        if not missing_reqs:
            return []
        
        # Try to find analysis for missing requirements in the raw response
        recovered_items = []
        
        # More aggressive patterns to find partial compliance items
        patterns = [
            # Pattern 1: Full compliance item
            r'"requirement"\s*:\s*"([^"]{20,})"[^}]*?"spec_value"\s*:\s*"([^"]*)"[^}]*?"complies"\s*:\s*(true|false)[^}]*?"justification"\s*:\s*"([^"]*)"',
            
            # Pattern 2: Partial items with just requirement and complies
            r'"requirement"\s*:\s*"([^"]{20,})"[^}]*?"complies"\s*:\s*(true|false)',
            
            # Pattern 3: Items that might be cut off
            r'(\d+)\.\s+([^"]{20,})"[^}]*?"complies"\s*:\s*(true|false)[^}]*?"justification"\s*:\s*"([^"]*)"',
        ]
        
        for pattern_idx, pattern in enumerate(patterns):
            print(f"🔧 Trying extraction pattern {pattern_idx + 1}")
            matches = re.findall(pattern, ai_content, re.DOTALL | re.IGNORECASE)
            
            for match in matches:
                if len(match) >= 2:  # At least requirement and complies
                    req_text = match[0] if pattern_idx < 2 else match[1]
                    normalized = re.sub(r'^\d+\.\s*', '', req_text).strip()[:50]
                    
                    # Check if this requirement is missing
                    if normalized not in analyzed_reqs:
                        item = {
                            "requirement": req_text.strip(),
                            "spec_value": match[1] if len(match) > 3 and pattern_idx == 0 else "Extracted from partial response",
                            "complies": match[-2].lower() == "true",  # Second to last element is usually complies
                            "justification": match[-1] if len(match) > 2 else "Partially extracted from AI response"
                        }
                        
                        recovered_items.append(item)
                        analyzed_reqs.add(normalized)
                        print(f"🔧 Recovered: {req_text[:50]}...")
        
        print(f"🔧 Recovered {len(recovered_items)} items from partial extraction")
        return recovered_items

    def export_to_excel(self, analysis_path: str, excel_path: str = None, override_meter: str = None, per_clause_override: dict = None) -> str:
        """Generate a detailed comparison report directly to Excel"""
        print(f"Starting Excel export for {analysis_path}...")
        sections = self._extract_sections_from_analysis(analysis_path)

        if not sections:
            print("No sections found in the analysis file.")
            return None

        # Create output file name if not specified
        if not excel_path:
            base_name = os.path.splitext(analysis_path)[0]
            excel_path = f"{base_name}_compliance_report.xlsx"

        # Create Excel workbook
        wb = Workbook()
        summary_sheet = wb.active
        summary_sheet.title = "Summary"

        # Add headers to summary sheet
        summary_headers = ['Clause', 'Meter', 'Overall Compliance', 'Compliant Items', 'Non-Compliant Items', 'Key Strengths']
        for i, header in enumerate(summary_headers, 1):
            summary_sheet.cell(row=1, column=i).value = header
            summary_sheet.cell(row=1, column=i).font = Font(bold=True)

        # Style for headers
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for col in range(1, len(summary_headers) + 1):
            summary_sheet.cell(row=1, column=col).fill = header_fill

        # Add data to summary sheet
        summary_row = 2
        all_sections_compliant = True

        for section_idx, section in enumerate(sections):
            clause_id = section.get('clause_id', 'Unknown')
            selected_meter = section.get('selected_meter')
            requirements = section.get('requirements', [])

            print(f"\nProcessing section {section_idx+1}/{len(sections)}: Clause {clause_id}")

            # Per-clause override takes precedence
            if per_clause_override and clause_id in per_clause_override:
                selected_meter = per_clause_override[clause_id]
                section['selected_meter'] = selected_meter
                print(f"  Using override meter: {selected_meter}")
            elif override_meter:
                selected_meter = override_meter
                section['selected_meter'] = override_meter
                print(f"  Using global override meter: {selected_meter}")
            else:
                print(f"  Using recommended meter: {selected_meter}")

            if not selected_meter:
                print(f"  No meter selected for clause {clause_id}, skipping")
                continue

            # Get meter specifications from DB
            meter_specs = self._find_meter_specs(selected_meter)

            # Update the meter description to match the selected meter
            section['meter_description'] = meter_specs.get('selection_blurb', '')

            if not meter_specs:
                all_sections_compliant = False
                print(f"❌ No specifications found for {selected_meter}")

                # Add to summary as non-compliant due to missing specs
                summary_sheet.cell(row=summary_row, column=1).value = clause_id
                summary_sheet.cell(row=summary_row, column=2).value = selected_meter
                summary_sheet.cell(row=summary_row, column=3).value = "❌ Non-compliant (specs not found)"
                summary_sheet.cell(row=summary_row, column=3).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                summary_row += 1
                continue

            # Compare requirements with specifications
            print(f"  Comparing {len(requirements)} requirements against {selected_meter} specifications...")
            comparison = self._safe_comparison_analysis(requirements, meter_specs, selected_meter)

            if "error" in comparison:
                all_sections_compliant = False
                print(f"❌ Error during comparison: {comparison['error']}")
                continue

            # Create detail sheet for each section
            sheet_name = f"Clause {clause_id}"
            if len(sheet_name) > 31:  # Excel sheet name length limit
                sheet_name = sheet_name[:31]
            detail_sheet = wb.create_sheet(title=sheet_name)

            # Add section info
            detail_sheet.cell(row=1, column=1).value = f"Clause {clause_id} Compliance Analysis"
            detail_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
            detail_sheet.cell(row=2, column=1).value = f"Selected Meter: {selected_meter}"
            detail_sheet.cell(row=2, column=1).font = Font(bold=True)
            detail_sheet.cell(row=3, column=1).value = f"Description: {section.get('meter_description', '')}"
            detail_sheet.cell(row=3, column=1).font = Font(italic=True)

            # Get analysis items
            analysis_items = comparison.get("compliance_analysis", [])
            print(f"  Got {len(analysis_items)} compliance analysis items")

            # Count compliant and non-compliant items
            compliant = 0
            non_compliant = 0
            for item in analysis_items:
                if item.get("complies", False):
                    compliant += 1
                else:
                    non_compliant += 1
                    all_sections_compliant = False

            # Overall compliance for this section
            overall_compliance = comparison.get("overall_compliance", False)
            status_text = "✅ Compliant" if overall_compliance else "❌ Non-compliant"
            detail_sheet.cell(row=4, column=1).value = f"Overall Compliance: {status_text}"
            print(f"  Overall compliance: {overall_compliance} ({compliant} compliant, {non_compliant} non-compliant)")

            # Add to summary
            summary_sheet.cell(row=summary_row, column=1).value = clause_id
            summary_sheet.cell(row=summary_row, column=2).value = selected_meter
            summary_sheet.cell(row=summary_row, column=3).value = status_text
            summary_sheet.cell(row=summary_row, column=4).value = compliant
            summary_sheet.cell(row=summary_row, column=5).value = non_compliant

            # Get areas exceeding requirements
            exceeding_areas = comparison.get("areas_exceeding_requirements", [])
            # Ensure all items are strings before joining
            exceeding_areas_str = [str(area) for area in exceeding_areas]
            summary_sheet.cell(row=summary_row, column=6).value = ', '.join(exceeding_areas_str)

            # Color code compliance
            if overall_compliance:
                summary_sheet.cell(row=summary_row, column=3).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                summary_sheet.cell(row=summary_row, column=3).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            summary_row += 1

            # Add table headers
            headers = ['Requirement', 'Specification', 'Status', 'Justification']
            for i, header in enumerate(headers, 1):
                detail_sheet.cell(row=6, column=i).value = header
                detail_sheet.cell(row=6, column=i).font = Font(bold=True)
                detail_sheet.cell(row=6, column=i).fill = header_fill

            # Add table data
            row_num = 7
            for item in analysis_items:
                # Format requirement
                req = item.get("requirement", "Unknown")
                detail_sheet.cell(row=row_num, column=1).value = req

                # Format specification
                spec = item.get("spec_value", "Not specified")
                detail_sheet.cell(row=row_num, column=2).value = spec

                # Format status with special handling for "exceeds"
                complies = item.get("complies", False)
                justification = item.get("justification", "")

                # Check if we have a better-than-required spec
                is_better = False
                if complies:
                    lower_justification = justification.lower()
                    if "exceed" in lower_justification or "better" in lower_justification or "surpass" in lower_justification:
                        is_better = True

                status_text = "✅ Exceeds" if is_better else "✅ Compliant" if complies else "❌ Non-compliant"
                detail_sheet.cell(row=row_num, column=3).value = status_text

                # Format justification
                detail_sheet.cell(row=row_num, column=4).value = justification

                # Color code status
                if complies:
                    detail_sheet.cell(row=row_num, column=3).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    detail_sheet.cell(row=row_num, column=3).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                row_num += 1

            # Add exceeding areas
            if exceeding_areas:
                row_num += 1
                detail_sheet.cell(row=row_num, column=1).value = "Areas Exceeding Requirements:"
                detail_sheet.cell(row=row_num, column=1).font = Font(bold=True)
                for area in exceeding_areas:
                    row_num += 1
                    detail_sheet.cell(row=row_num, column=1).value = f"- {area}"

            # Add issues
            issues = comparison.get("potential_issues", [])
            if issues:
                row_num += 1
                detail_sheet.cell(row=row_num, column=1).value = "Potential Compliance Issues:"
                detail_sheet.cell(row=row_num, column=1).font = Font(bold=True)
                for issue in issues:
                    row_num += 1
                    detail_sheet.cell(row=row_num, column=1).value = f"- {issue}"

            # Auto-adjust column widths
            for col in range(1, 5):
                max_length = 0
                for row in range(6, row_num + 1):
                    cell_value = detail_sheet.cell(row=row, column=col).value
                    if cell_value:
                        text_length = len(str(cell_value).split('\n')[0])  # First line length
                        max_length = max(max_length, min(text_length, 80))  # Cap at 80

                # Set width with some padding
                adjusted_width = max_length + 2
                detail_sheet.column_dimensions[get_column_letter(col)].width = adjusted_width

            # Add borders and text wrapping
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))

            for row in range(6, row_num):
                for col in range(1, 5):
                    detail_sheet.cell(row=row, column=col).border = thin_border
                    detail_sheet.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')

        # Auto-adjust summary column widths
        for col in range(1, len(summary_headers) + 1):
            max_length = len(str(summary_headers[col-1]))
            for row in range(2, summary_row):
                cell_value = summary_sheet.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Set width with some padding
            summary_sheet.column_dimensions[get_column_letter(col)].width = max_length + 2

        # Add borders to summary
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

        for row in range(1, summary_row):
            for col in range(1, len(summary_headers) + 1):
                summary_sheet.cell(row=row, column=col).border = thin_border

        # Save workbook
        wb.save(excel_path)
        print(f"✅ Excel report saved to: {excel_path}")
        return excel_path

    def generate_detailed_comparison(self, analysis_path: str, override_meter: str = None, per_clause_override: dict = None) -> str:
        """Generate a detailed comparison report with properly formatted tables"""
        sections = self._extract_sections_from_analysis(analysis_path)

        if not sections:
            return "No sections found in the analysis file."

        # Create output file name
        base_name = os.path.splitext(analysis_path)[0]
        output_path = f"{base_name}_detailed_comparison.md"

        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("# DETAILED METER SPECIFICATION COMPLIANCE ANALYSIS\n\n")

            for section in sections:
                clause_id = section.get('clause_id', 'Unknown')
                requirements = section.get('requirements', [])
                selected_meter = section.get('selected_meter')

                # Per-clause override takes precedence
                if per_clause_override and clause_id in per_clause_override:
                    selected_meter = per_clause_override[clause_id]
                    section['selected_meter'] = selected_meter
                elif override_meter:
                    selected_meter = override_meter
                    section['selected_meter'] = override_meter

                f.write(f"## Clause {clause_id} Compliance Analysis\n\n")

                if not selected_meter:
                    f.write("No meter selected for this clause.\n\n")
                    continue

                f.write(f"### Selected Meter: {selected_meter}\n\n")
                f.write(f"**Description**: {section.get('meter_description', '')}\n\n")

                # Get meter specifications from DB
                meter_specs = self._find_meter_specs(selected_meter)

                if not meter_specs:
                    f.write(f"⚠️ No detailed specifications found for {selected_meter} in the database.\n\n")
                    continue

                print(f"Analyzing compliance for {selected_meter} against {clause_id} requirements...")

                # Compare requirements with specifications
                comparison = self._safe_comparison_analysis(requirements, meter_specs, selected_meter)

                # Write comparison results
                f.write("### Detailed Compliance Analysis\n\n")

                if "error" in comparison:
                    f.write(f"⚠️ Error analyzing compliance: {comparison['error']}\n\n")
                    continue

                # Get analysis items
                analysis_items = comparison.get("compliance_analysis", [])

                if not analysis_items:
                    f.write("No compliance analysis items generated.\n\n")
                    continue

                # Create properly formatted table with consistent widths
                f.write("| Requirement | Specification | Status | Justification |\n")
                f.write("|------------|---------------|--------|---------------|\n")

                for item in analysis_items:
                    req = item.get("requirement", "Unknown").strip()
                    spec = item.get("spec_value", "Not specified").strip()

                    # Fix the compliance display for better-than-required specs
                    complies = item.get("complies", False)
                    justification = item.get("justification", "").strip()

                    # Check if we have a better-than-required spec that was marked compliant
                    is_better = False
                    if complies:
                        lower_justification = justification.lower()
                        if "exceed" in lower_justification or "better" in lower_justification:
                            is_better = True

                    compliance_status = "✅ Exceeds" if is_better else "✅ Compliant" if complies else "❌ Non-compliant"

                    # Format and write row
                    req = self._format_cell_content(req, 30)
                    spec = self._format_cell_content(spec, 30)
                    justification = self._format_cell_content(justification, 40)

                    f.write(f"| {req} | {spec} | {compliance_status} | {justification} |\n")

                f.write("\n")

                # Overall assessment
                overall = comparison.get("overall_compliance", False)
                f.write(f"**Overall Compliance**: {'✅ Compliant' if overall else '❌ Non-compliant'}\n\n")

                # Areas exceeding requirements
                exceeding = comparison.get("areas_exceeding_requirements", [])
                if exceeding:
                    f.write("**Areas Exceeding Requirements**:\n")
                    for area in exceeding:
                        f.write(f"- {area}\n")
                    f.write("\n")

                # Potential issues
                issues = comparison.get("potential_issues", [])
                if issues:
                    f.write("**Potential Compliance Issues**:\n")
                    for issue in issues:
                        f.write(f"- {issue}\n")
                    f.write("\n")

                f.write("\n" + "-"*80 + "\n\n")

            f.write("\n*Analysis completed using AI-assisted specification comparison*\n")

        print(f"✅ Detailed comparison saved to: {output_path}")
        return output_path

    def _format_cell_content(self, text: str, max_length: int = 30) -> str:
        """Format cell content for better table readability"""
        if not text or len(text) <= max_length:
            return text

        # Split into words
        words = text.split()
        lines = []
        current_line = []
        current_length = 0

        for word in words:
            # If adding this word exceeds the max length, start a new line
            if current_length + len(word) + (1 if current_length > 0 else 0) > max_length:
                lines.append(" ".join(current_line))
                current_line = [word]
                current_length = len(word)
            else:
                # Add to current line
                current_line.append(word)
                current_length += len(word) + (1 if current_length > 0 else 0)

        # Add the last line if not empty
        if current_line:
            lines.append(" ".join(current_line))

        # Join with <br> for markdown line breaks in table cells
        return "<br>".join(lines)

    def _extract_and_repair_json(self, text: str) -> Dict:
        """Extract JSON from text and repair common issues with Qwen output"""
        # First, try to find complete JSON inside the response
        json_pattern = re.compile(r'\{(?:[^{}]|(?:\{[^{}]*\}))*\}', re.DOTALL)
        matches = json_pattern.findall(text)
        
        # Try each match, from longest to shortest (assuming longer is more complete)
        if matches:
            matches.sort(key=len, reverse=True)
            for match in matches:
                try:
                    return json.loads(match)
                except json.JSONDecodeError:
                    pass
    
        # If no valid JSON found with regex, try more aggressive cleaning
        clean_text = text
    
        # Remove any text before the first '{'
        start_idx = clean_text.find('{')
        if start_idx >= 0:
            clean_text = clean_text[start_idx:]
    
        # Remove any text after the last '}'
        end_idx = clean_text.rfind('}')
        if end_idx >= 0:
            clean_text = clean_text[:end_idx+1]
    
        # Fix common JSON errors
        try:
            # Fix unquoted property names (the most common issue)
            clean_text = re.sub(r'([{,]\s*)([a-zA-Z_][a-zA-Z0-9_]*)\s*:', r'\1"\2":', clean_text)
            
            # Fix single quotes used instead of double quotes
            clean_text = re.sub(r"'([^']*)'", r'"\1"', clean_text)
            
            # Fix boolean capitalization
            clean_text = re.sub(r':\s*True\b', r': true', clean_text)
            clean_text = re.sub(r':\s*False\b', r': false', clean_text)
            
            # Remove trailing commas in objects and arrays
            clean_text = re.sub(r',\s*}', '}', clean_text)
            clean_text = re.sub(r',\s*]', ']', clean_text)
            
            # Try to parse the fixed JSON
            return json.loads(clean_text)
        except json.JSONDecodeError as e:
            print(f"⚠️ Standard JSON repair failed: {e}")
        
        # Last resort: manual extraction of data
        print("Attempting manual extraction of JSON data...")
        try:
            result = {"compliance_analysis": [], "overall_compliance": False}
            
            # Find compliance_analysis section
            analysis_match = re.search(r'"compliance_analysis"\s*:\s*\[(.*?)\]', text, re.DOTALL)
            if analysis_match:
                analysis_text = analysis_match.group(1)
                
                # Extract individual items
                item_pattern = re.compile(r'{(.*?)}', re.DOTALL)
                items = item_pattern.findall(analysis_text)
                
                for item_text in items:
                    item = {}
                    
                    # Extract requirement
                    req_match = re.search(r'"requirement"\s*:\s*"([^"]*)"', item_text)
                    if req_match:
                        item["requirement"] = req_match.group(1)
                    
                    # Extract spec_value
                    spec_match = re.search(r'"spec_value"\s*:\s*"([^"]*)"', item_text)
                    if spec_match:
                        item["spec_value"] = spec_match.group(1)
                    
                    # Extract complies
                    complies_match = re.search(r'"complies"\s*:\s*(true|false)', item_text.lower())
                    if complies_match:
                        item["complies"] = complies_match.group(1) == "true"
                    
                    # Extract justification
                    just_match = re.search(r'"justification"\s*:\s*"([^"]*)"', item_text)
                    if just_match:
                        item["justification"] = just_match.group(1)
                    
                    if "requirement" in item:
                        result["compliance_analysis"].append(item)
            
            # If we extracted any items, consider it a success
            if result["compliance_analysis"]:
                return result
        except Exception as e:
            print(f"Manual extraction failed: {e}")
        
        # If everything fails, return a synthetic response
        print("Creating fallback JSON response...")
        return {
            "compliance_analysis": [],
            "overall_compliance": False,
            "areas_exceeding_requirements": [],
            "potential_issues": ["Could not parse AI response due to JSON formatting errors"]
        }
    def _post_process_compliance_logic(self, result: dict, original_requirements: list) -> dict:
        """Correct compliance logic for accuracy classes and percentages."""
        import re
        analysis_items = result.get("compliance_analysis", [])
        corrected_count = 0
        exceeding_areas = []

        for item in analysis_items:
            req = item.get("requirement", "").lower()
            spec = item.get("spec_value", "").lower()
            justification = item.get("justification", "").lower()
            original_complies = item.get("complies", False)

            # --- Accuracy Class Correction ---
            req_class = re.search(r'class\s*([\d\.]+)\s*s?', req)
            spec_class = re.search(r'class\s*([\d\.]+)\s*s?', spec)
            if req_class and spec_class:
                req_val = float(req_class.group(1))
                spec_val = float(spec_class.group(1))
                # If spec is better (smaller) or equal, mark as compliant
                if spec_val <= req_val and not original_complies:
                    item["complies"] = True
                    item["justification"] = (
                        f"CORRECTED: Meter class {spec_class.group(0)} is better than required {req_class.group(0)} (smaller class number = higher accuracy)."
                    )
                    exceeding_areas.append(f"{item.get('requirement','')} ({item.get('spec_value','')})")
                    corrected_count += 1

            # --- Percentage Accuracy Correction ---
            req_pct = re.search(r'±?(\d+\.?\d*)\s*%', req)
            spec_pct = re.search(r'±?(\d+\.?\d*)\s*%', spec)
            if req_pct and spec_pct:
                req_val = float(req_pct.group(1))
                spec_val = float(spec_pct.group(1))
                if spec_val <= req_val and not original_complies:
                    item["complies"] = True
                    item["justification"] = (
                        f"CORRECTED: Meter accuracy ±{spec_val}% is better than required ±{req_val}% (smaller percentage = higher accuracy)."
                    )
                    exceeding_areas.append(f"{item.get('requirement','')} ({item.get('spec_value','')})")
                    corrected_count += 1

            # --- Justification Phrase Correction ---
            if (
                ("more stringent than" in justification or
                "exceeds requirement" in justification or
                "better than" in justification or
                "higher accuracy" in justification or
                "surpasses" in justification)
                and not original_complies
            ):
                item["complies"] = True
                item["justification"] = f"CORRECTED: {item['justification']} (meter exceeds requirement)"
                exceeding_areas.append(f"{item.get('requirement','')} ({item.get('spec_value','')})")
                corrected_count += 1

        if corrected_count > 0:
            print(f"🔧 Corrected {corrected_count} compliance logic errors (meter exceeds requirement)")
            result["areas_exceeding_requirements"] = list(set(result.get("areas_exceeding_requirements", []) + exceeding_areas))
            # Recalculate overall compliance
            total = len(analysis_items)
            compliant = sum(1 for item in analysis_items if item.get("complies", False))
            result["overall_compliance"] = (compliant / total) >= 0.8 if total else False

        return result
    
    def _clean_json_output(self, content):
        """Clean up JSON output from Qwen for better parsing"""
        
        # Qwen sometimes adds markdown code fences, remove them
        content = re.sub(r'```json\s*', '', content)
        content = re.sub(r'```\s*$', '', content)
        
        # Handle the case where Qwen adds explanations before or after JSON
        json_match = re.search(r'(\{.*\})', content, re.DOTALL)
        if json_match:
            content = json_match.group(1)
        
        # Fix common Qwen JSON issues
        content = content.replace('true,\n}', 'true\n}')
        content = content.replace('false,\n}', 'false\n}')
        content = content.replace("'", '"')
        content = re.sub(r',\s*]', ']', content)
        content = re.sub(r',\s*}', '}', content)
        
        return content
    
    def _format_meter_specs_for_prompt(self, meter_specs: Dict) -> str:
        """Format meter specifications in a concise way for the AI prompt"""
        lines = []
        
        # Basic meter info
        if "model_name" in meter_specs and meter_specs["model_name"]:
            lines.append(f"Model: {meter_specs['model_name']}")
        if "product_name" in meter_specs and meter_specs["product_name"]:
            lines.append(f"Product: {meter_specs['product_name']}")
        if "series_name" in meter_specs and meter_specs["series_name"]:
            lines.append(f"Series: {meter_specs['series_name']}")
        if "selection_blurb" in meter_specs and meter_specs["selection_blurb"]:
            lines.append(f"Description: {meter_specs['selection_blurb']}")
        
        # Physical specifications
        lines.append("\nPHYSICAL SPECIFICATIONS:")
        if "display_type" in meter_specs and meter_specs["display_type"]:
            lines.append(f"- Display: {meter_specs['display_type']}")
        if "mounting_mode" in meter_specs and meter_specs["mounting_mode"]:
            lines.append(f"- Mounting: {meter_specs['mounting_mode']}")
        if "rated_current" in meter_specs and meter_specs["rated_current"]:
            lines.append(f"- Rated Current: {meter_specs['rated_current']}")
        if "network_frequency" in meter_specs and meter_specs["network_frequency"]:
            lines.append(f"- Frequency: {meter_specs['network_frequency']}")
        if "sampling_rate" in meter_specs and meter_specs["sampling_rate"]:
            lines.append(f"- Sampling Rate: {meter_specs['sampling_rate']}")
        if "memory_capacity" in meter_specs and meter_specs["memory_capacity"]:
            lines.append(f"- Memory: {meter_specs['memory_capacity']}")
        
        # Environmental specifications
        if meter_specs.get("operating_temp") or meter_specs.get("storage_temp"):
            lines.append("\nENVIRONMENTAL SPECIFICATIONS:")
            if "operating_temp" in meter_specs and meter_specs["operating_temp"]:
                lines.append(f"- Operating Temperature: {meter_specs['operating_temp']}")
            if "storage_temp" in meter_specs and meter_specs["storage_temp"]:
                lines.append(f"- Storage Temperature: {meter_specs['storage_temp']}")
            if "relative_humidity" in meter_specs and meter_specs["relative_humidity"]:
                lines.append(f"- Humidity: {meter_specs['relative_humidity']}")
                
        # Accuracy specifications (most important for compliance)
        if "measurement_accuracy" in meter_specs and meter_specs["measurement_accuracy"]:
            lines.append("\nMEASUREMENT ACCURACY:")
            for param, value in meter_specs["measurement_accuracy"].items():
                lines.append(f"- {param}: {value}")
                
        # Accuracy classes
        if "accuracy_classes" in meter_specs and meter_specs["accuracy_classes"]:
            lines.append("\nACCURACY CLASSES:")
            for acc_class in meter_specs["accuracy_classes"]:
                lines.append(f"- {acc_class}")
                
        # Communication protocols
        if "communication_protocols" in meter_specs and meter_specs["communication_protocols"]:
            lines.append("\nCOMMUNICATION PROTOCOLS:")
            for protocol, support in meter_specs["communication_protocols"].items():
                support_text = f" ({support})" if support else ""
                lines.append(f"- {protocol}{support_text}")
                    
        # Power quality features
        if "power_quality_features" in meter_specs and meter_specs["power_quality_features"]:
            lines.append("\nPOWER QUALITY ANALYSIS:")
            for feature in meter_specs["power_quality_features"]:
                lines.append(f"- {feature}")
                    
        # Measurements
        if "measurements" in meter_specs and meter_specs["measurements"]:
            lines.append("\nMEASUREMENT CAPABILITIES:")
            for measurement in meter_specs["measurements"]:
                lines.append(f"- {measurement}")
                
        # Data recording
        if "data_recording" in meter_specs and meter_specs["data_recording"]:
            lines.append("\nDATA RECORDING:")
            for recording in meter_specs["data_recording"]:
                lines.append(f"- {recording}")
                
        # Inputs/Outputs
        if "inputs_outputs" in meter_specs and meter_specs["inputs_outputs"]:
            lines.append("\nINPUTS/OUTPUTS:")
            for io in meter_specs["inputs_outputs"]:
                lines.append(f"- {io['type']}: {io['description']}")
                
        # Certifications
        if "certifications" in meter_specs and meter_specs["certifications"]:
            lines.append("\nCERTIFICATIONS:")
            for cert in meter_specs["certifications"]:
                lines.append(f"- {cert}")
                
        # Applications
        if "applications" in meter_specs and meter_specs["applications"]:
            lines.append("\nAPPLICATIONS:")
            for app in meter_specs["applications"]:
                lines.append(f"- {app}")
                
        return "\n".join(lines)
    
    def _safe_remove_duplicates(self, items: list) -> list:
        """Safely remove duplicates from a list that might contain unhashable types"""
        if not items:
            return []
        
        unique_items = []
        seen = set()
        
        for item in items:
            try:
                # Try to use the item as a set member (hashable types)
                if item not in seen:
                    seen.add(item)
                    unique_items.append(item)
            except TypeError:
                # If unhashable, check manually
                if item not in unique_items:
                    unique_items.append(item)
        
        return unique_items

    def _safe_comparison_analysis(self, requirements: List[str], meter_specs: Dict, model_number: str) -> Dict:
        """Wrapper around comparison with better error handling"""
        try:
            print(f"🔧 DEBUG: Starting comparison for {model_number}")
            print(f"🔧 DEBUG: Requirements type: {type(requirements)}, count: {len(requirements) if requirements else 0}")
            print(f"🔧 DEBUG: Meter specs type: {type(meter_specs)}, keys: {list(meter_specs.keys()) if meter_specs else 'None'}")
            
            # Check for problematic data types in requirements
            for i, req in enumerate(requirements):
                if not isinstance(req, str):
                    print(f"🔧 DEBUG: WARNING - Requirement {i} is not a string: {type(req)}")
            
            # Check for problematic data types in meter_specs
            for key, value in meter_specs.items():
                if isinstance(value, dict):
                    for sub_key, sub_value in value.items():
                        if not isinstance(sub_key, (str, int, float)):
                            print(f"🔧 DEBUG: WARNING - Meter spec dict key is not hashable: {type(sub_key)}")
            
            result = self._compare_requirements_with_specs(requirements, meter_specs, model_number)
            
            print(f"🔧 DEBUG: Comparison completed successfully")
            return result
            
        except TypeError as e:
            print(f"🔧 DEBUG: TypeError in comparison: {e}")
            print(f"🔧 DEBUG: Error details: {str(e)}")
            import traceback
            traceback.print_exc()
            
            # Return a safe fallback
            return {
                "error": f"TypeError during comparison: {str(e)}",
                "compliance_analysis": [],
                "overall_compliance": False,
                "areas_exceeding_requirements": [],
                "potential_issues": []
            }
            
        except Exception as e:
            print(f"🔧 DEBUG: Other error in comparison: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            return {
                "error": f"Comparison failed: {str(e)}",
                "compliance_analysis": [],
                "overall_compliance": False,
                "areas_exceeding_requirements": [],
                "potential_issues": []
            }
       

if __name__ == "__main__":
    print("\nMeter Specification Compliance Tool\n" + "="*35)

    # Get analysis file path
    analysis_file = input("Enter path to meter analysis file (.txt): ").strip('"\'')
    if not os.path.exists(analysis_file):
        print(f"Error: Analysis file '{analysis_file}' not found!")
        exit(1)

    # Create comparator
    comparator = MeterSpecificationComparison()

    # Extract sections first
    sections = comparator._extract_sections_from_analysis(analysis_file)
    if not sections:
        print("No sections found in the analysis file.")
        exit(1)

    # Collect per-clause overrides
    per_clause_override = {}
    print("\nReview recommended meters for each clause. Press Enter to accept or type an alternate model number.\n")
    for section in sections:
        clause_id = section.get('clause_id', 'Unknown')
        recommended = section.get('selected_meter', 'None')
        print(f"Clause {clause_id} - {recommended}")
        alt = input("Alternate Meter (Enter to default): ").strip()
        if alt:
            per_clause_override[clause_id] = alt
        else:
            per_clause_override[clause_id] = recommended

    # Ask for output format preference
    output_format = input("\nSelect output format:\n1. Markdown\n2. Excel\n3. Both\nEnter choice (1-3): ").strip()

    # Add indicator here
    print(f"[INFO] Output format selected: {output_format}. Proceeding with report generation...")

    # Generate outputs based on preference
    md_output = None
    excel_output = None

    try: 
        if output_format in ["1", "3"]:
            print("[INFO] Generating Markdown report...")
            md_output = comparator.generate_detailed_comparison(
                analysis_file,
                override_meter=None,
                per_clause_override=per_clause_override
            )
            print(f"Markdown report generated: {md_output}")

        if output_format in ["2", "3"]:
            print("[INFO] Generating Excel report...")
            excel_output = comparator.export_to_excel(
                analysis_file,
                override_meter=None,
                per_clause_override=per_clause_override
            )
            print(f"Excel report generated: {excel_output}")

        print("\nAnalysis complete!")
        
    except Exception as e:
        print(f"\n❌ ERROR: {type(e).__name__}: {e}")
        print("The process encountered an error. Please check your input files and database.")